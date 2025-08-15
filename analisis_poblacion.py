#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Análisis de Población por Grupos Etarios y Ubicación Geográfica
Autor: Sistema de Análisis Demográfico
Descripción: Script para analizar población por municipio, vereda y grupo etario
             desde un CSV sin encabezados, generando reportes en Excel.
"""

# ============================================
# IMPORTACIÓN DE LIBRERÍAS
# ============================================
import pandas as pd
import numpy as np
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import warnings
import os
import sys

warnings.filterwarnings('ignore')

# Configurar pandas para evitar problemas con Excel
pd.options.mode.chained_assignment = None  # Para evitar warnings
pd.set_option('display.max_columns', None)  # Para debug si es necesario

# ============================================
# CONFIGURACIÓN
# ============================================

COLUMNA_MUNICIPIO = 2        # Posición de la columna municipio
COLUMNA_VEREDA = 8          # Posición de la columna vereda  
COLUMNA_FECHA_REGISTRO = 4   # Posición columna fecha registro (formato: 2021-08-27 14:54:39.243)
COLUMNA_FECHA_NACIMIENTO = 18 # Posición columna fecha nacimiento (formato: 1947-11-16)

# Nombre del archivo CSV a procesar
ARCHIVO_CSV = 'poblacion_veredas.csv'

# Orden estándar de grupos etarios para todo el reporte
ORDEN_GRUPOS_ETARIOS = ['Menor de 9 meses', '09-23 meses', '02-19 años', '20-59 años', '60+ años', 'Sin datos']

# ============================================
# FUNCIONES DE PROCESAMIENTO
# ============================================

def cargar_datos(archivo_csv):
    """Carga el CSV sin encabezados"""
    print(f"\n{'='*60}")
    print("CARGANDO DATOS")
    print('='*60)
    
    try:
        df = pd.read_csv(archivo_csv, header=None)
        print(f"✅ CSV cargado exitosamente: {len(df):,} registros")
        print(f"📊 Número de columnas: {len(df.columns)}")
        
        # Mostrar verificación de columnas
        print("\n--- VERIFICACIÓN DE COLUMNAS SELECCIONADAS ---")
        if len(df) > 0:
            print(f"Columna {COLUMNA_MUNICIPIO} (Municipio): {df.iloc[0, COLUMNA_MUNICIPIO]}")
            print(f"Columna {COLUMNA_VEREDA} (Vereda): {df.iloc[0, COLUMNA_VEREDA]}")
            print(f"Columna {COLUMNA_FECHA_REGISTRO} (Fecha Registro): {df.iloc[0, COLUMNA_FECHA_REGISTRO]}")
            print(f"Columna {COLUMNA_FECHA_NACIMIENTO} (Fecha Nacimiento): {df.iloc[0, COLUMNA_FECHA_NACIMIENTO]}")
            print("\n¿Son correctas estas columnas? Si no, ajusta las posiciones en la configuración.")
        
        return df
    except FileNotFoundError:
        print(f"❌ ERROR: No se encontró el archivo '{archivo_csv}'")
        print("Verifica que el archivo esté en el mismo directorio que este script")
        sys.exit(1)
    except Exception as e:
        print(f"❌ ERROR al cargar el archivo: {e}")
        sys.exit(1)

def convertir_fechas(df):
    """Convierte las columnas de fecha al formato correcto"""
    print(f"\n{'='*60}")
    print("CONVIRTIENDO FECHAS")
    print('='*60)
    
    # Convertir fecha de registro (incluye timestamp con milisegundos)
    try:
        df[COLUMNA_FECHA_REGISTRO] = pd.to_datetime(df[COLUMNA_FECHA_REGISTRO], 
                                                    format='%Y-%m-%d %H:%M:%S.%f', 
                                                    errors='coerce')
        # Si hay muchos NaT, intenta formato sin milisegundos
        if df[COLUMNA_FECHA_REGISTRO].isna().sum() > len(df) * 0.5:
            df[COLUMNA_FECHA_REGISTRO] = pd.to_datetime(df[COLUMNA_FECHA_REGISTRO], 
                                                        format='%Y-%m-%d %H:%M:%S', 
                                                        errors='coerce')
    except:
        df[COLUMNA_FECHA_REGISTRO] = pd.to_datetime(df[COLUMNA_FECHA_REGISTRO], 
                                                    errors='coerce')
    
    # Convertir fecha de nacimiento (solo fecha)
    df[COLUMNA_FECHA_NACIMIENTO] = pd.to_datetime(df[COLUMNA_FECHA_NACIMIENTO], 
                                                  format='%Y-%m-%d', 
                                                  errors='coerce')
    
    # Verificar fechas nulas
    fechas_registro_nulas = df[COLUMNA_FECHA_REGISTRO].isna().sum()
    fechas_nacimiento_nulas = df[COLUMNA_FECHA_NACIMIENTO].isna().sum()
    
    print(f"📅 Fechas de registro nulas/inválidas: {fechas_registro_nulas:,}")
    print(f"📅 Fechas de nacimiento nulas/inválidas: {fechas_nacimiento_nulas:,}")
    
    if fechas_registro_nulas > 0 or fechas_nacimiento_nulas > 0:
        print("\n⚠️ Advertencia: Hay fechas nulas que serán excluidas del análisis")
    
    # Eliminar registros con fechas nulas
    df_limpio = df.dropna(subset=[COLUMNA_FECHA_REGISTRO, COLUMNA_FECHA_NACIMIENTO])
    print(f"\n✅ Registros válidos para análisis: {len(df_limpio):,}")
    
    # Validación de fechas lógicas
    fechas_invalidas = df_limpio[df_limpio[COLUMNA_FECHA_NACIMIENTO] > df_limpio[COLUMNA_FECHA_REGISTRO]]
    if len(fechas_invalidas) > 0:
        print(f"\n⚠️ Advertencia: {len(fechas_invalidas)} registros tienen fecha de nacimiento posterior a fecha de registro")
        df_limpio = df_limpio[df_limpio[COLUMNA_FECHA_NACIMIENTO] <= df_limpio[COLUMNA_FECHA_REGISTRO]]
        print(f"Registros válidos después de filtrar fechas ilógicas: {len(df_limpio):,}")
    
    return df_limpio, df

def calcular_edad_detallada(fecha_nacimiento, fecha_referencia):
    """Calcula la edad en años y meses totales"""
    if pd.isna(fecha_nacimiento) or pd.isna(fecha_referencia):
        return None, None
    
    if fecha_nacimiento > fecha_referencia:
        return None, None
    
    diferencia = relativedelta(fecha_referencia, fecha_nacimiento)
    edad_anos = diferencia.years
    edad_meses_total = diferencia.years * 12 + diferencia.months
    
    return edad_anos, edad_meses_total

def calcular_edades(df_limpio):
    """Calcula las edades y valida los datos"""
    print(f"\n{'='*60}")
    print("CALCULANDO EDADES Y VALIDANDO DATOS")
    print('='*60)
    
    fecha_actual = pd.Timestamp.now()
    print(f"📅 Fecha de referencia para cálculo de edad: {fecha_actual.strftime('%Y-%m-%d')}")
    
    # Calcular edades usando fecha actual
    edades = df_limpio.apply(lambda row: calcular_edad_detallada(
        row[COLUMNA_FECHA_NACIMIENTO], 
        fecha_actual  # CAMBIO: Usar fecha actual en lugar de fecha de registro
    ), axis=1)
    
    df_limpio['edad_anos'] = edades.apply(lambda x: x[0] if x else None)
    df_limpio['edad_meses'] = edades.apply(lambda x: x[1] if x else None)
    
    # Calcular edad en días usando fecha actual
    df_limpio['edad_dias'] = (fecha_actual - df_limpio[COLUMNA_FECHA_NACIMIENTO]).dt.days
    
    # Agregar columna de días desde registro (útil para análisis temporal)
    df_limpio['dias_desde_registro'] = (fecha_actual - df_limpio[COLUMNA_FECHA_REGISTRO]).dt.days
    
    # Detectar anomalías (sin incluir bebés como sospechosos)
    print("\n📊 VALIDACIÓN DE CALIDAD DE DATOS")
    print("-" * 40)
    
    # Información sobre bebés (NO sospechoso, solo informativo)
    menores_1_mes = df_limpio[df_limpio['edad_dias'] < 30]
    if len(menores_1_mes) > 0:
        print(f"\n👶 Bebés menores de 1 mes: {len(menores_1_mes)} (esto es normal)")
        print(f"  - Menos de 7 días: {len(df_limpio[df_limpio['edad_dias'] < 7])}")
        print(f"  - Entre 7-14 días: {len(df_limpio[(df_limpio['edad_dias'] >= 7) & (df_limpio['edad_dias'] < 14)])}")
        print(f"  - Entre 15-30 días: {len(df_limpio[(df_limpio['edad_dias'] >= 15) & (df_limpio['edad_dias'] < 30)])}")
    
    # Mayores de 100 años (SÍ sospechoso)
    mayores_100 = df_limpio[df_limpio['edad_anos'] > 100]
    if len(mayores_100) > 0:
        print(f"\n⚠️ ALERTA: Personas mayores de 100 años: {len(mayores_100)}")
        edades_mayores = mayores_100['edad_anos'].sort_values(ascending=False).head()
        print(f"  Edades encontradas: {edades_mayores.tolist()}")
        print("  Estos casos requieren verificación")
    
    # Verificar si hay edades negativas (error grave)
    edades_negativas = df_limpio[df_limpio['edad_dias'] < 0]
    if len(edades_negativas) > 0:
        print(f"\n❌ ERROR GRAVE: {len(edades_negativas)} registros con edad negativa")
        print("  Esto indica fechas de nacimiento futuras - revisar datos")
    
    # Estadísticas finales
    print(f"\n📊 ESTADÍSTICAS DE EDAD (calculadas al {fecha_actual.strftime('%Y-%m-%d')}):")
    print(f"  Edad mínima: {df_limpio['edad_dias'].min()} días ({df_limpio['edad_anos'].min():.1f} años)")
    print(f"  Edad máxima: {df_limpio['edad_dias'].max():,} días ({df_limpio['edad_anos'].max():.1f} años)")
    print(f"  Edad promedio: {df_limpio['edad_anos'].mean():.1f} años")
    print(f"  Edad mediana: {df_limpio['edad_anos'].median():.1f} años")
    
    # Información sobre antigüedad de registros
    print(f"\n📅 ANTIGÜEDAD DE LOS REGISTROS:")
    print(f"  Registro más reciente: hace {df_limpio['dias_desde_registro'].min()} días")
    print(f"  Registro más antiguo: hace {df_limpio['dias_desde_registro'].max()} días")
    print(f"  Promedio: hace {df_limpio['dias_desde_registro'].mean():.0f} días")
    
    return df_limpio

def clasificar_grupo_etario(edad_meses):
    """Clasifica según los grupos etarios especificados"""
    if pd.isna(edad_meses):
        return 'Sin datos'
    
    if edad_meses < 9:
        return 'Menor de 9 meses'
    elif edad_meses >= 9 and edad_meses <= 23:
        return '09-23 meses'
    elif edad_meses >= 24 and edad_meses < 240:  # 2 a 19 años
        return '02-19 años'
    elif edad_meses >= 240 and edad_meses < 720:  # 20 a 59 años
        return '20-59 años'
    else:
        return '60+ años'

def asignar_grupos_etarios(df_limpio):
    """Asigna grupos etarios y nombres a las columnas"""
    print(f"\n{'='*60}")
    print("ASIGNANDO GRUPOS ETARIOS")
    print('='*60)
    
    # Asignar grupo etario
    df_limpio['grupo_etario'] = df_limpio['edad_meses'].apply(clasificar_grupo_etario)
    
    # Asignar nombres a las columnas
    df_limpio['municipio'] = df_limpio.iloc[:, COLUMNA_MUNICIPIO]
    df_limpio['vereda'] = df_limpio.iloc[:, COLUMNA_VEREDA]
    
    # Mostrar distribución
    print("\nDistribución por grupo etario:")
    for grupo in ORDEN_GRUPOS_ETARIOS:
        if grupo in df_limpio['grupo_etario'].value_counts().index:
            cantidad = df_limpio['grupo_etario'].value_counts()[grupo]
            print(f"  {grupo:20} : {cantidad:,} personas")
    
    print(f"\n📍 Municipios únicos: {df_limpio['municipio'].nunique()}")
    print(f"📍 Veredas únicas: {df_limpio['vereda'].nunique()}")
    
    return df_limpio

def crear_tablas_conteo(df_limpio):
    """Crea las tablas de conteo por vereda y municipio"""
    print(f"\n{'='*60}")
    print("CREANDO TABLAS DE CONTEO")
    print('='*60)
    
    # Conteo detallado
    conteo_detallado = df_limpio.groupby(['municipio', 'vereda', 'grupo_etario']).size().reset_index(name='cantidad')
    
    # Tabla pivot para veredas
    tabla_pivot = conteo_detallado.pivot_table(
        index=['municipio', 'vereda'],
        columns='grupo_etario',
        values='cantidad',
        fill_value=0,
        aggfunc='sum'
    ).reset_index()
    
    # Ordenar columnas
    columnas_ordenadas = ['municipio', 'vereda']
    for grupo in ORDEN_GRUPOS_ETARIOS:
        if grupo in tabla_pivot.columns:
            columnas_ordenadas.append(grupo)
    
    columnas_grupos = [col for col in tabla_pivot.columns if col not in ['municipio', 'vereda']]
    tabla_pivot['Total'] = tabla_pivot[columnas_grupos].sum(axis=1)
    columnas_ordenadas.append('Total')
    tabla_pivot = tabla_pivot[columnas_ordenadas]
    
    # Tabla por municipios
    totales_municipio = df_limpio.groupby(['municipio', 'grupo_etario']).size().reset_index(name='cantidad')
    tabla_municipios = totales_municipio.pivot_table(
        index='municipio',
        columns='grupo_etario',
        values='cantidad',
        fill_value=0,
        aggfunc='sum'
    ).reset_index()
    
    # Ordenar columnas municipios
    columnas_ordenadas_mun = ['municipio']
    for grupo in ORDEN_GRUPOS_ETARIOS:
        if grupo in tabla_municipios.columns:
            columnas_ordenadas_mun.append(grupo)
    columnas_ordenadas_mun.append('Total')
    
    columnas_grupos_mun = [col for col in tabla_municipios.columns if col != 'municipio']
    tabla_municipios['Total'] = tabla_municipios[columnas_grupos_mun].sum(axis=1)
    tabla_municipios = tabla_municipios[columnas_ordenadas_mun]
    
    print(f"✅ Tabla de veredas creada: {len(tabla_pivot)} registros")
    print(f"✅ Tabla de municipios creada: {len(tabla_municipios)} registros")
    
    return tabla_pivot, tabla_municipios, conteo_detallado

def limpiar_para_excel(valor):
    """Limpia valores para evitar problemas con Excel"""
    if pd.isna(valor):
        return ''
    valor_str = str(valor)
    # Si empieza con =, +, -, @ (caracteres que Excel interpreta como fórmulas)
    if valor_str and valor_str[0] in ['=', '+', '-', '@']:
        return "'" + valor_str  # Agregar comilla simple para forzar como texto
    return valor_str

def crear_resumen_general(df_limpio, df_original):
    """Crea el resumen general detallado"""
    print(f"\n{'='*60}")
    print("CREANDO RESUMEN GENERAL")
    print('='*60)
    
    # Fecha actual para referencia
    fecha_actual = pd.Timestamp.now()
    
    # Calcular estadísticas
    edad_minima_dias = df_limpio['edad_dias'].min()
    edad_maxima_dias = df_limpio['edad_dias'].max()
    edad_minima_meses = df_limpio['edad_meses'].min()
    edad_maxima_meses = df_limpio['edad_meses'].max()
    edad_minima_anos = df_limpio['edad_anos'].min()
    edad_maxima_anos = df_limpio['edad_anos'].max()
    
    # Conteo por grupos
    conteo_grupos = df_limpio['grupo_etario'].value_counts().sort_index()
    total_personas = len(df_limpio)
    
    # Casos especiales (no necesariamente sospechosos)
    bebes_menores_30_dias = len(df_limpio[df_limpio['edad_dias'] < 30])
    mayores_100 = len(df_limpio[df_limpio['edad_anos'] > 100])
    edades_negativas = len(df_limpio[df_limpio['edad_dias'] < 0])
    
    # Categorías especiales
    menores_9_meses = len(df_limpio[df_limpio['edad_meses'] < 9])
    menores_1_ano = len(df_limpio[df_limpio['edad_meses'] < 12])
    menores_2_anos = len(df_limpio[df_limpio['edad_meses'] < 24])
    menores_5_anos = len(df_limpio[df_limpio['edad_meses'] < 60])
    menores_18_anos = len(df_limpio[df_limpio['edad_meses'] < 216])
    adultos_mayores_70 = len(df_limpio[df_limpio['edad_meses'] >= 840])
    adultos_mayores_80 = len(df_limpio[df_limpio['edad_meses'] >= 960])
    
    # Indicadores demográficos
    poblacion_edad_escolar = len(df_limpio[(df_limpio['edad_anos'] >= 5) & (df_limpio['edad_anos'] <= 17)])
    poblacion_edad_laboral = len(df_limpio[(df_limpio['edad_anos'] >= 18) & (df_limpio['edad_anos'] <= 65)])
    
    # Crear DataFrame de resumen
    filas_resumen = []
    
    # Información general
    filas_resumen.extend([
        {'Categoría': '>>> INFORMACIÓN GENERAL', 'Métrica': '', 'Valor': ''},
        {'Categoría': '', 'Métrica': 'Fecha y hora del reporte', 'Valor': fecha_actual.strftime('%Y-%m-%d %H:%M:%S')},
        {'Categoría': '', 'Métrica': 'Fecha de cálculo de edades', 'Valor': fecha_actual.strftime('%Y-%m-%d')},
        {'Categoría': '', 'Métrica': 'Total de registros en CSV original', 'Valor': f"{len(df_original):,}"},
        {'Categoría': '', 'Métrica': 'Registros con fechas válidas', 'Valor': f"{len(df_limpio):,}"},
        {'Categoría': '', 'Métrica': 'Registros excluidos', 'Valor': f"{len(df_original) - len(df_limpio):,}"},
        {'Categoría': '', 'Métrica': 'Porcentaje de datos válidos', 'Valor': f"{(len(df_limpio)/len(df_original)*100):.1f}%"},
        {'Categoría': '', 'Métrica': '', 'Valor': ''},
    ])
    
    # Información demográfica especial
    filas_resumen.extend([
        {'Categoría': '>>> CASOS ESPECIALES', 'Métrica': '', 'Valor': ''},
        {'Categoría': '', 'Métrica': 'Bebés menores de 30 días', 'Valor': f"{bebes_menores_30_dias:,} (normal en registros)"},
        {'Categoría': '', 'Métrica': '⚠️ Personas mayores de 100 años', 'Valor': f"{mayores_100:,} (requiere verificación)"},
        {'Categoría': '', 'Métrica': '❌ Edades negativas (error)', 'Valor': f"{edades_negativas:,}"},
        {'Categoría': '', 'Métrica': '', 'Valor': ''},
    ])
    
    # Cobertura geográfica
    filas_resumen.extend([
        {'Categoría': '>>> COBERTURA GEOGRÁFICA', 'Métrica': '', 'Valor': ''},
        {'Categoría': '', 'Métrica': 'Total de municipios', 'Valor': f"{df_limpio['municipio'].nunique():,}"},
        {'Categoría': '', 'Métrica': 'Total de veredas', 'Valor': f"{df_limpio['vereda'].nunique():,}"},
        {'Categoría': '', 'Métrica': 'Promedio personas por vereda', 'Valor': f"{len(df_limpio) / df_limpio['vereda'].nunique():.1f}"},
        {'Categoría': '', 'Métrica': '', 'Valor': ''},
    ])
    
    # Estadísticas de edad
    filas_resumen.extend([
        {'Categoría': '>>> ESTADÍSTICAS DE EDAD (A FECHA ACTUAL)', 'Métrica': '', 'Valor': ''},
        {'Categoría': '', 'Métrica': 'Edad mínima', 'Valor': f"{int(edad_minima_dias):,} días / {int(edad_minima_meses)} meses / {edad_minima_anos:.1f} años"},
        {'Categoría': '', 'Métrica': 'Edad máxima', 'Valor': f"{int(edad_maxima_dias):,} días / {int(edad_maxima_meses):,} meses / {edad_maxima_anos:.1f} años"},
        {'Categoría': '', 'Métrica': 'Edad promedio', 'Valor': f"{df_limpio['edad_anos'].mean():.1f} años"},
        {'Categoría': '', 'Métrica': 'Edad mediana', 'Valor': f"{df_limpio['edad_anos'].median():.1f} años"},
        {'Categoría': '', 'Métrica': 'Desviación estándar', 'Valor': f"{df_limpio['edad_anos'].std():.1f} años"},
        {'Categoría': '', 'Métrica': '', 'Valor': ''},
    ])
    
    # Distribución por grupo etario
    filas_resumen.append({'Categoría': '>>> DISTRIBUCIÓN POR GRUPO ETARIO', 'Métrica': '', 'Valor': ''})
    for grupo in ORDEN_GRUPOS_ETARIOS:
        if grupo in conteo_grupos.index:
            cantidad = conteo_grupos[grupo]
            porcentaje = (cantidad / total_personas) * 100
            filas_resumen.append({'Categoría': '', 'Métrica': grupo, 'Valor': f"{cantidad:,} personas ({porcentaje:.1f}%)"})
        else:
            filas_resumen.append({'Categoría': '', 'Métrica': grupo, 'Valor': "0 personas (0.0%)"})
    filas_resumen.extend([
        {'Categoría': '', 'Métrica': 'TOTAL', 'Valor': f"{total_personas:,} personas (100.0%)"},
        {'Categoría': '', 'Métrica': '', 'Valor': ''},
    ])
    
    # Categorías especiales de interés
    filas_resumen.extend([
        {'Categoría': '>>> CATEGORÍAS DE INTERÉS POBLACIONAL', 'Métrica': '', 'Valor': ''},
        {'Categoría': '', 'Métrica': 'Bebés < 9 meses', 'Valor': f"{menores_9_meses:,} ({menores_9_meses/total_personas*100:.1f}%)"},
        {'Categoría': '', 'Métrica': 'Infantes < 1 año', 'Valor': f"{menores_1_ano:,} ({menores_1_ano/total_personas*100:.1f}%)"},
        {'Categoría': '', 'Métrica': 'Primera infancia < 5 años', 'Valor': f"{menores_5_anos:,} ({menores_5_anos/total_personas*100:.1f}%)"},
        {'Categoría': '', 'Métrica': 'Menores de edad < 18 años', 'Valor': f"{menores_18_anos:,} ({menores_18_anos/total_personas*100:.1f}%)"},
        {'Categoría': '', 'Métrica': 'Adultos mayores 70+ años', 'Valor': f"{adultos_mayores_70:,} ({adultos_mayores_70/total_personas*100:.1f}%)"},
        {'Categoría': '', 'Métrica': 'Adultos mayores 80+ años', 'Valor': f"{adultos_mayores_80:,} ({adultos_mayores_80/total_personas*100:.1f}%)"},
        {'Categoría': '', 'Métrica': '', 'Valor': ''},
    ])
    
    # Antigüedad de registros
    if 'dias_desde_registro' in df_limpio.columns:
        promedio_dias_registro = df_limpio['dias_desde_registro'].mean()
        min_dias_registro = df_limpio['dias_desde_registro'].min()
        max_dias_registro = df_limpio['dias_desde_registro'].max()
        
        filas_resumen.extend([
            {'Categoría': '>>> ANTIGÜEDAD DE REGISTROS', 'Métrica': '', 'Valor': ''},
            {'Categoría': '', 'Métrica': 'Registro más reciente', 'Valor': f"hace {min_dias_registro} días"},
            {'Categoría': '', 'Métrica': 'Registro más antiguo', 'Valor': f"hace {max_dias_registro} días"},
            {'Categoría': '', 'Métrica': 'Promedio antigüedad', 'Valor': f"hace {promedio_dias_registro:.0f} días"},
        ])
    
    resumen_general = pd.DataFrame(filas_resumen)
    
    print("✅ Resumen general creado con éxito")
    
    return resumen_general

def identificar_casos_sospechosos(df_limpio):
    """Identifica casos sospechosos para revisión"""
    print(f"\n{'='*60}")
    print("IDENTIFICANDO CASOS SOSPECHOSOS")
    print('='*60)
    
    casos_sospechosos = pd.DataFrame()
    
    # NO incluir bebés menores de 30 días - es normal tener recién nacidos
    
    # Personas mayores de 100 años (SÍ son sospechosos)
    centenarios = df_limpio[df_limpio['edad_anos'] > 100].copy()
    if len(centenarios) > 0:
        centenarios['tipo_caso'] = 'Mayor de 100 años'
        centenarios['observacion'] = centenarios['edad_anos'].apply(lambda x: f'{x:.0f} años')
        casos_sospechosos = pd.concat([casos_sospechosos, centenarios])
    
    # Edades negativas (error grave)
    edades_negativas = df_limpio[df_limpio['edad_dias'] < 0].copy()
    if len(edades_negativas) > 0:
        edades_negativas['tipo_caso'] = 'Edad negativa (fecha nacimiento futura)'
        edades_negativas['observacion'] = 'ERROR: Fecha de nacimiento posterior a hoy'
        casos_sospechosos = pd.concat([casos_sospechosos, edades_negativas])
    
    # Registros muy antiguos con bebés (posible error)
    # Si el registro es de hace más de 2 años pero la persona tiene menos de 2 años
    if 'dias_desde_registro' in df_limpio.columns:
        registros_antiguos_bebes = df_limpio[
            (df_limpio['dias_desde_registro'] > 730) &  # Registro de hace más de 2 años
            (df_limpio['edad_anos'] < 2)  # Pero la persona tiene menos de 2 años
        ].copy()
        if len(registros_antiguos_bebes) > 0:
            registros_antiguos_bebes['tipo_caso'] = 'Inconsistencia temporal'
            registros_antiguos_bebes['observacion'] = 'Bebé con registro muy antiguo'
            casos_sospechosos = pd.concat([casos_sospechosos, registros_antiguos_bebes])
    
    # Eliminar duplicados
    if len(casos_sospechosos) > 0:
        casos_sospechosos = casos_sospechosos.drop_duplicates(subset=[COLUMNA_FECHA_NACIMIENTO, COLUMNA_FECHA_REGISTRO])
        
        # Columnas base para el reporte
        columnas_reporte = ['municipio', 'vereda', COLUMNA_FECHA_NACIMIENTO, COLUMNA_FECHA_REGISTRO, 
                           'edad_dias', 'edad_meses', 'edad_anos']
        
        # Agregar días desde registro si existe
        if 'dias_desde_registro' in casos_sospechosos.columns:
            columnas_reporte.append('dias_desde_registro')
        
        columnas_reporte.extend(['tipo_caso', 'observacion'])
        
        casos_sospechosos_reporte = casos_sospechosos[columnas_reporte].copy()
        
        # Renombrar columnas
        nuevos_nombres = {
            'municipio': 'Municipio',
            'vereda': 'Vereda',
            COLUMNA_FECHA_NACIMIENTO: 'Fecha_Nacimiento',
            COLUMNA_FECHA_REGISTRO: 'Fecha_Registro',
            'edad_dias': 'Edad_Dias',
            'edad_meses': 'Edad_Meses',
            'edad_anos': 'Edad_Años',
            'dias_desde_registro': 'Días_Desde_Registro',
            'tipo_caso': 'Tipo_Caso',
            'observacion': 'Observación'
        }
        casos_sospechosos_reporte.rename(columns=nuevos_nombres, inplace=True)
        
        print(f"⚠️ Total de casos sospechosos encontrados: {len(casos_sospechosos_reporte)}")
        print("\nResumen por tipo:")
        for tipo in casos_sospechosos['tipo_caso'].unique():
            cantidad = len(casos_sospechosos[casos_sospechosos['tipo_caso'] == tipo])
            print(f"  - {tipo}: {cantidad}")
    else:
        casos_sospechosos_reporte = pd.DataFrame()
        print("✅ No se encontraron casos sospechosos")
    
    return casos_sospechosos_reporte

def guardar_excel(tabla_pivot, tabla_municipios, resumen_general, conteo_detallado, casos_sospechosos_reporte):
    """Guarda los resultados en un archivo Excel formateado"""
    print(f"\n{'='*60}")
    print("GUARDANDO ARCHIVO EXCEL")
    print('='*60)
    
    # Generar nombre del archivo
    fecha_reporte = datetime.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f'reporte_poblacion_grupos_etarios_{fecha_reporte}.xlsx'
    
    try:
        # Crear el archivo Excel
        with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
            # Guardar hojas
            tabla_pivot.to_excel(writer, sheet_name='Detalle_por_Vereda', index=False)
            tabla_municipios.to_excel(writer, sheet_name='Resumen_por_Municipio', index=False)
            
            # Para la hoja de estadísticas, asegurar que no haya problemas con fórmulas
            # Convertir todos los valores a string para evitar interpretación de fórmulas
            resumen_general_safe = resumen_general.copy()
            for col in resumen_general_safe.columns:
                resumen_general_safe[col] = resumen_general_safe[col].astype(str)
            resumen_general_safe.to_excel(writer, sheet_name='Estadisticas_Generales', index=False)
            
            conteo_detallado.to_excel(writer, sheet_name='Datos_Conteo_Detallado', index=False)
            
            if len(casos_sospechosos_reporte) > 0:
                # Renombrar la hoja sin caracteres especiales al inicio
                casos_sospechosos_reporte.to_excel(writer, sheet_name='Casos_Sospechosos', index=False)
            
            # Ajustar anchos de columna
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            cell_value = str(cell.value) if cell.value else ''
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Aplicar formato a encabezados
            from openpyxl.styles import PatternFill, Font, Alignment
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            alert_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            
            # Formatear encabezados de cada hoja
            for sheet_name in ['Detalle_por_Vereda', 'Resumen_por_Municipio']:
                if sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
            
            # Formato especial para casos sospechosos
            if 'Casos_Sospechosos' in writer.sheets:
                ws = writer.sheets['Casos_Sospechosos']
                for cell in ws[1]:
                    cell.fill = alert_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            
            # Aplicar formato especial a la hoja de estadísticas
            if 'Estadisticas_Generales' in writer.sheets:
                ws = writer.sheets['Estadisticas_Generales']
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Hacer las filas de categoría en negrita
                for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row[0].value and '>>>' in str(row[0].value):
                        row[0].font = Font(bold=True, size=11)
        
        print(f"✅ Archivo guardado exitosamente: {nombre_archivo}")
        print(f"\n📋 El archivo contiene {4 if len(casos_sospechosos_reporte) == 0 else 5} hojas:")
        print("  1. Detalle_por_Vereda")
        print("  2. Resumen_por_Municipio")
        print("  3. Estadisticas_Generales")
        print("  4. Datos_Conteo_Detallado")
        if len(casos_sospechosos_reporte) > 0:
            print(f"  5. Casos_Sospechosos ({len(casos_sospechosos_reporte)} registros)")
        
        return nombre_archivo
    
    except Exception as e:
        print(f"❌ ERROR al guardar el archivo Excel: {e}")
        return None

def mostrar_visualizacion_final(df_limpio):
    """Muestra una visualización final de la distribución"""
    print(f"\n{'='*60}")
    print("DISTRIBUCIÓN PORCENTUAL POR GRUPO ETARIO")
    print('='*60)
    
    total_personas = len(df_limpio)
    conteo_grupos = df_limpio['grupo_etario'].value_counts().sort_index()
    
    for grupo in conteo_grupos.index:
        cantidad = conteo_grupos[grupo]
        porcentaje = (cantidad / total_personas) * 100
        barra = '█' * int(porcentaje / 2)
        print(f"{grupo:20} {cantidad:6,} personas ({porcentaje:5.1f}%) {barra}")

def main():
    """Función principal que ejecuta todo el análisis"""
    print("\n" + "="*80)
    print(" ANÁLISIS DE POBLACIÓN POR GRUPOS ETARIOS Y UBICACIÓN GEOGRÁFICA ".center(80))
    print("="*80)
    fecha_actual = datetime.now()
    print(f"Iniciando análisis: {fecha_actual.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"📌 NOTA: Las edades se calculan a fecha ACTUAL ({fecha_actual.strftime('%Y-%m-%d')})")
    print("         NO a la fecha de registro del CSV")
    print("-"*80)
    
    try:
        # 1. Cargar datos
        df_original = cargar_datos(ARCHIVO_CSV)
        
        # 2. Convertir fechas y limpiar datos
        df_limpio, df_original = convertir_fechas(df_original)
        
        # 3. Calcular edades
        df_limpio = calcular_edades(df_limpio)
        
        # 4. Asignar grupos etarios
        df_limpio = asignar_grupos_etarios(df_limpio)
        
        # 5. Crear tablas de conteo
        tabla_pivot, tabla_municipios, conteo_detallado = crear_tablas_conteo(df_limpio)
        
        # 6. Crear resumen general
        resumen_general = crear_resumen_general(df_limpio, df_original)
        
        # 7. Identificar casos sospechosos
        casos_sospechosos = identificar_casos_sospechosos(df_limpio)
        
        # 8. Guardar en Excel
        archivo_excel = guardar_excel(tabla_pivot, tabla_municipios, resumen_general, 
                                     conteo_detallado, casos_sospechosos)
        
        # 9. Mostrar visualización final
        mostrar_visualizacion_final(df_limpio)
        
        # 10. Resumen final
        print(f"\n{'='*80}")
        print(" ANÁLISIS COMPLETADO EXITOSAMENTE ".center(80))
        print('='*80)
        
        if archivo_excel:
            print(f"\n📊 Archivo de resultados: {archivo_excel}")
            print(f"📁 Ubicación: {os.path.abspath(archivo_excel)}")
        
        print(f"\n⏱️ Finalizado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("\n¡Análisis completado con éxito! Revisa el archivo Excel generado.")
        
    except Exception as e:
        print(f"\n❌ ERROR GENERAL: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# ============================================
# EJECUCIÓN PRINCIPAL
# ============================================
if __name__ == "__main__":
    # Verificar dependencias
    try:
        import openpyxl
    except ImportError:
        print("❌ ERROR: La librería 'openpyxl' no está instalada.")
        print("Instálala ejecutando: pip install openpyxl")
        sys.exit(1)
    
    main()