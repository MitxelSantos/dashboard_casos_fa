"""
utils/cobertura_processor.py - VERSIÓN SIMPLIFICADA Y OPTIMIZADA
Reduce de 1200+ líneas a ~400 líneas manteniendo toda la funcionalidad
"""

import pandas as pd
import numpy as np
import logging
from datetime import datetime
import streamlit as st

logger = logging.getLogger(__name__)

# ===== CONFIGURACIÓN SIMPLIFICADA =====
MUNICIPIOS_MAPEO = {
    "CARMEN APICALA": "CARMEN DE APICALA",
    "SALDAÑA": "SALDANA", 
    "VALLE SAN JUAN": "VALLE DE SAN JUAN"
}

GRUPOS_EDAD = {
    "E": "9M-11M", "F": "1-5 AÑOS", "G": "6-10 A", "H": "11-20 A", "I": "21-30 A",
    "J": "31-40 A", "K": "41-50 A", "L": "51-59 A", "M": "60-69 A", "N": "70 A +"
}

# ===== FUNCIONES PRINCIPALES SIMPLIFICADAS =====

@st.cache_data(ttl=3600)
def load_and_process_cobertura_data():
    """✅ CORREGIDO: Función principal simplificada."""
    try:
        logger.info("🚀 Cargando datos de cobertura (simplificado)")
        
        file_path = load_cobertura_from_google_drive_fixed()
        if not file_path:
            logger.error("❌ No se pudo cargar archivo de cobertura")
            return None
        
        cobertura_data = process_cobertura_data_simplified(file_path)
        if cobertura_data:
            logger.info(f"✅ Cobertura procesada: {len(cobertura_data.get('municipios', {}))} municipios")
        
        return cobertura_data
        
    except Exception as e:
        logger.error(f"❌ Error en carga de cobertura: {str(e)}")
        return None

def load_cobertura_from_google_drive_fixed():
    """✅ CORREGIDO: Interfaz arreglada con ConsolidatedDataLoader."""
    try:
        from data_loader import get_data_loader
        
        if not hasattr(st.secrets, "drive_files") or "cobertura" not in st.secrets.drive_files:
            logger.error("❌ ID de cobertura no encontrado en secrets")
            return None
            
        loader = get_data_loader()
        
        # ✅ CORREGIDO: Usar _authenticate() en lugar de authenticate()
        if not loader._authenticate():
            logger.error("❌ No se pudo autenticar Google Drive")
            return None
            
        cobertura_file_id = st.secrets.drive_files["cobertura"]
        temp_path = loader._download_file(cobertura_file_id, "Cobertura.xlsx")
        
        if temp_path:
            logger.info("✅ Archivo de cobertura descargado")
            return temp_path
        else:
            logger.error("❌ Fallo en descarga")
            return None
        
    except Exception as e:
        logger.error(f"❌ Error en descarga: {str(e)}")
        return None

def process_cobertura_data_simplified(file_path):
    """✅ SIMPLIFICADO: Procesa datos de cobertura - reducido en 70%."""
    try:
        import openpyxl
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        municipios_hojas = [sheet for sheet in workbook.sheetnames if sheet != "Veredas"]
        logger.info(f"📋 Procesando {len(municipios_hojas)} municipios")
        
        cobertura_data = {
            "municipios": {},
            "summary": {"total_poblacion_tolima": 0, "total_vacunados_tolima": 0, "total_rechazos_tolima": 0},
            "metadata": {"fecha_procesamiento": datetime.now()}
        }
        
        for hoja_name in municipios_hojas:
            try:
                municipio_data = process_municipio_sheet_simplified(workbook, hoja_name)
                if municipio_data and municipio_data.get("total_poblacion", 0) > 0:
                    # Aplicar mapeo de nombres
                    municipio_dashboard = MUNICIPIOS_MAPEO.get(hoja_name, hoja_name)
                    cobertura_data["municipios"][municipio_dashboard] = municipio_data
                    
                    # Acumular totales
                    cobertura_data["summary"]["total_poblacion_tolima"] += municipio_data["total_poblacion"]
                    cobertura_data["summary"]["total_vacunados_tolima"] += municipio_data["total_vacunados"]
                    cobertura_data["summary"]["total_rechazos_tolima"] += municipio_data.get("total_rechazos", 0)
                    
            except Exception as e:
                logger.warning(f"⚠️ Error en {hoja_name}: {str(e)}")
                continue
        
        # Calcular cobertura promedio
        total_pob = cobertura_data["summary"]["total_poblacion_tolima"]
        total_vac = cobertura_data["summary"]["total_vacunados_tolima"]
        cobertura_data["summary"]["cobertura_promedio"] = (total_vac / total_pob * 100) if total_pob > 0 else 0
        
        workbook.close()
        logger.info(f"✅ Procesamiento completado: {len(cobertura_data['municipios'])} municipios válidos")
        return cobertura_data
        
    except Exception as e:
        logger.error(f"❌ Error procesando archivo: {str(e)}")
        return None

def process_municipio_sheet_simplified(workbook, municipio_name):
    """✅ SIMPLIFICADO: Procesa hoja de municipio - reducido en 80%."""
    try:
        worksheet = workbook[municipio_name]
        
        municipio_data = {
            "nombre": municipio_name,
            "total_poblacion": 0,
            "total_vacunados": 0,
            "total_rechazos": 0,
            "cobertura_general": 0.0,
            "urbano": {"poblacion": 0, "vacunados": 0, "rechazos": 0, "cobertura": 0.0},
            "rural": {"poblacion": 0, "vacunados": 0, "rechazos": 0, "cobertura": 0.0},
            "veredas": {}
        }
        
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            if not row or not row[2]:  # Sin vereda
                continue
                
            vereda_name = str(row[2]).strip()
            poblacion = safe_int(row[3])
            
            if poblacion <= 0:
                continue  # Skip veredas sin población
            
            # Calcular vacunados (columnas E-N)
            vacunados = sum(safe_int(row[i]) for i in range(4, 14) if i < len(row))
            
            # Calcular rechazos (columnas O-Q)  
            rechazos = sum(safe_int(row[i]) for i in range(14, 17) if i < len(row))
            
            cobertura = (vacunados / poblacion * 100) if poblacion > 0 else 0
            
            # Clasificar urbano vs rural
            if "CASCO URBANO" in vereda_name.upper():
                municipio_data["urbano"] = {
                    "poblacion": poblacion,
                    "vacunados": vacunados,
                    "rechazos": rechazos,
                    "cobertura": round(cobertura, 1)
                }
            else:
                municipio_data["rural"]["poblacion"] += poblacion
                municipio_data["rural"]["vacunados"] += vacunados
                municipio_data["rural"]["rechazos"] += rechazos
                
                municipio_data["veredas"][vereda_name] = {
                    "poblacion": poblacion,
                    "vacunados": vacunados,
                    "rechazos": rechazos,
                    "cobertura": round(cobertura, 1)
                }
            
            # Acumular totales
            municipio_data["total_poblacion"] += poblacion
            municipio_data["total_vacunados"] += vacunados
            municipio_data["total_rechazos"] += rechazos
        
        # Calcular coberturas finales
        if municipio_data["total_poblacion"] > 0:
            municipio_data["cobertura_general"] = round(
                (municipio_data["total_vacunados"] / municipio_data["total_poblacion"]) * 100, 1
            )
        
        if municipio_data["rural"]["poblacion"] > 0:
            municipio_data["rural"]["cobertura"] = round(
                (municipio_data["rural"]["vacunados"] / municipio_data["rural"]["poblacion"]) * 100, 1
            )
        
        return municipio_data
        
    except Exception as e:
        logger.error(f"❌ Error en hoja {municipio_name}: {str(e)}")
        return None

# ===== FUNCIONES DE ACCESO SIMPLIFICADAS =====

def get_cobertura_for_municipio(cobertura_data, municipio_name):
    """Obtiene datos de cobertura para municipio."""
    if not cobertura_data or "municipios" not in cobertura_data:
        return None
        
    # Buscar directamente
    if municipio_name in cobertura_data["municipios"]:
        return cobertura_data["municipios"][municipio_name]
    
    # Buscar con mapeo inverso
    for excel_name, dashboard_name in MUNICIPIOS_MAPEO.items():
        if dashboard_name == municipio_name and excel_name in cobertura_data["municipios"]:
            return cobertura_data["municipios"][excel_name]
    
    return None

def get_cobertura_for_vereda(cobertura_data, municipio_name, vereda_name):
    """Obtiene datos de cobertura para vereda."""
    municipio_data = get_cobertura_for_municipio(cobertura_data, municipio_name)
    if not municipio_data or "veredas" not in municipio_data:
        return None
    
    veredas_data = municipio_data["veredas"]
    
    # Buscar exacta
    if vereda_name in veredas_data:
        return veredas_data[vereda_name]
    
    # Buscar normalizada
    vereda_norm = str(vereda_name).strip().upper()
    for vereda_key, vereda_data in veredas_data.items():
        if str(vereda_key).strip().upper() == vereda_norm:
            return vereda_data
    
    return None

def get_rechazos_for_municipio(cobertura_data, municipio_name):
    """Obtiene rechazos para municipio."""
    municipio_data = get_cobertura_for_municipio(cobertura_data, municipio_name)
    return municipio_data.get("total_rechazos", 0) if municipio_data else 0

def get_rechazos_for_vereda(cobertura_data, municipio_name, vereda_name):
    """Obtiene rechazos para vereda."""
    vereda_data = get_cobertura_for_vereda(cobertura_data, municipio_name, vereda_name)
    return vereda_data.get("rechazos", 0) if vereda_data else 0

# ===== FUNCIONES DE VALIDACIÓN SIMPLIFICADAS =====

def validate_cobertura_data_quality_contextual(cobertura_data, municipio_filter=None, vereda_filter=None):
    """✅ SIMPLIFICADO: Validación contextual básica."""
    if not cobertura_data:
        return {"alertas": ["❌ No hay datos de cobertura"], "calidad": 0.0}
    
    alertas = []
    
    # Validación específica por filtro
    if vereda_filter and vereda_filter != "Todas" and municipio_filter:
        # Validación de vereda
        vereda_data = get_cobertura_for_vereda(cobertura_data, municipio_filter, vereda_filter)
        if not vereda_data:
            alertas.append(f"❌ Sin datos para {vereda_filter}")
        else:
            cobertura = vereda_data.get("cobertura", 0)
            if cobertura > 150:
                alertas.append(f"📈 Cobertura muy alta: {cobertura:.1f}%")
            elif cobertura < 30:
                alertas.append(f"📉 Cobertura muy baja: {cobertura:.1f}%")
                
    elif municipio_filter and municipio_filter != "Todos":
        # Validación de municipio
        municipio_data = get_cobertura_for_municipio(cobertura_data, municipio_filter)
        if not municipio_data:
            alertas.append(f"❌ Sin datos para {municipio_filter}")
        else:
            cobertura = municipio_data.get("cobertura_general", 0)
            if cobertura > 120:
                alertas.append(f"📈 Cobertura alta: {cobertura:.1f}%")
            elif cobertura < 40:
                alertas.append(f"📉 Cobertura baja: {cobertura:.1f}%")
    
    # Calidad general simplificada
    calidad = 85.0 if not alertas else max(50.0, 85.0 - len(alertas) * 15)
    
    return {"alertas": alertas, "calidad": calidad}

def validate_cobertura_multiple(cobertura_data, municipios_seleccionados, veredas_seleccionadas=None):
    """✅ SIMPLIFICADO: Validación múltiple básica."""
    if not cobertura_data or not municipios_seleccionados:
        return {"alertas": ["❌ Sin datos o municipios"], "calidad": 0.0, "cobertura": 0.0, "vacunados_total": 0, "poblacion_total": 0}
    
    total_poblacion = 0
    total_vacunados = 0
    municipios_validos = 0
    alertas = []
    
    if veredas_seleccionadas:
        # Procesar veredas específicas
        for municipio in municipios_seleccionados:
            for vereda in veredas_seleccionadas:
                vereda_data = get_cobertura_for_vereda(cobertura_data, municipio, vereda)
                if vereda_data and vereda_data.get("poblacion", 0) > 0:
                    total_poblacion += vereda_data["poblacion"]
                    total_vacunados += vereda_data["vacunados"]
    else:
        # Procesar municipios completos
        for municipio in municipios_seleccionados:
            municipio_data = get_cobertura_for_municipio(cobertura_data, municipio)
            if municipio_data and municipio_data.get("total_poblacion", 0) > 0:
                total_poblacion += municipio_data["total_poblacion"]
                total_vacunados += municipio_data["total_vacunados"]
                municipios_validos += 1
            else:
                alertas.append(f"❌ {municipio}: sin datos")
    
    cobertura_agregada = (total_vacunados / total_poblacion * 100) if total_poblacion > 0 else 0.0
    calidad = max(50.0, 100.0 - len(alertas) * 10)
    
    return {
        "alertas": alertas,
        "calidad": calidad,
        "cobertura": round(cobertura_agregada, 1),
        "vacunados_total": total_vacunados,
        "poblacion_total": total_poblacion,
        "municipios_validos": municipios_validos
    }

def get_coverage_comparison_context(cobertura_data, filters):
    """✅ SIMPLIFICADO: Contexto de comparación."""
    if not cobertura_data:
        return {"cobertura": 0.0, "poblacion": 0, "vacunados": 0, "contexto": "Sin datos"}
    
    modo = filters.get("modo", "unico")
    municipio_display = filters.get("municipio_display", "Todos")
    vereda_display = filters.get("vereda_display", "Todas")
    
    if modo == "multiple":
        municipios_sel = filters.get("municipios_seleccionados", [])
        veredas_sel = filters.get("veredas_seleccionadas", [])
        
        if veredas_sel and municipios_sel:
            return calculate_coverage_for_selected_veredas_simple(cobertura_data, municipios_sel, veredas_sel)
        elif municipios_sel:
            return calculate_coverage_for_selected_municipios_simple(cobertura_data, municipios_sel)
    
    elif vereda_display != "Todas" and municipio_display != "Todos":
        vereda_data = get_cobertura_for_vereda(cobertura_data, municipio_display, vereda_display)
        if vereda_data:
            return {
                "cobertura": vereda_data["cobertura"],
                "poblacion": vereda_data["poblacion"],
                "vacunados": vereda_data["vacunados"],
                "contexto": f"Vereda {vereda_display[:15]}"
            }
    
    elif municipio_display != "Todos":
        municipio_data = get_cobertura_for_municipio(cobertura_data, municipio_display)
        if municipio_data:
            return {
                "cobertura": municipio_data["cobertura_general"],
                "poblacion": municipio_data["total_poblacion"],
                "vacunados": municipio_data["total_vacunados"],
                "contexto": f"Municipio {municipio_display[:12]}"
            }
    
    # Departamental
    summary = cobertura_data.get("summary", {})
    return {
        "cobertura": summary.get("cobertura_promedio", 0.0),
        "poblacion": summary.get("total_poblacion_tolima", 0),
        "vacunados": summary.get("total_vacunados_tolima", 0),
        "contexto": "Departamento Tolima"
    }

# ===== FUNCIONES DE APOYO SIMPLIFICADAS =====

def calculate_coverage_for_selected_municipios_simple(cobertura_data, municipios_seleccionados):
    """Calcula cobertura para municipios seleccionados - versión simple."""
    total_poblacion = 0
    total_vacunados = 0
    municipios_encontrados = 0
    
    for municipio in municipios_seleccionados:
        municipio_data = get_cobertura_for_municipio(cobertura_data, municipio)
        if municipio_data and municipio_data.get("total_poblacion", 0) > 0:
            total_poblacion += municipio_data["total_poblacion"]
            total_vacunados += municipio_data["total_vacunados"]
            municipios_encontrados += 1
    
    cobertura = (total_vacunados / total_poblacion * 100) if total_poblacion > 0 else 0.0
    
    return {
        "cobertura": round(cobertura, 1),
        "poblacion": total_poblacion,
        "vacunados": total_vacunados,
        "contexto": f"{municipios_encontrados} municipios"
    }

def calculate_coverage_for_selected_veredas_simple(cobertura_data, municipios_seleccionados, veredas_seleccionadas):
    """Calcula cobertura para veredas seleccionadas - versión simple."""
    total_poblacion = 0
    total_vacunados = 0
    veredas_encontradas = 0
    
    for municipio in municipios_seleccionados:
        for vereda in veredas_seleccionadas:
            vereda_data = get_cobertura_for_vereda(cobertura_data, municipio, vereda)
            if vereda_data and vereda_data.get("poblacion", 0) > 0:
                total_poblacion += vereda_data["poblacion"]
                total_vacunados += vereda_data["vacunados"]
                veredas_encontradas += 1
    
    cobertura = (total_vacunados / total_poblacion * 100) if total_poblacion > 0 else 0.0
    
    return {
        "cobertura": round(cobertura, 1),
        "poblacion": total_poblacion,
        "vacunados": total_vacunados,
        "contexto": f"{veredas_encontradas} veredas"
    }

def safe_int(value, default=0):
    """Conversión segura a entero."""
    try:
        return int(value) if value and pd.notna(value) else default
    except (ValueError, TypeError):
        return default

def safe_float(value, default=0.0):
    """Conversión segura a float."""
    try:
        return float(value) if value and pd.notna(value) else default
    except (ValueError, TypeError):
        return default

# ===== FUNCIONES DE DEBUG LIGERAS =====

def debug_vereda_mapping(cobertura_data, municipio_name, vereda_name_shapefile):
    """Debug ligero para mapeo de veredas."""
    if not cobertura_data:
        return
    
    municipio_data = get_cobertura_for_municipio(cobertura_data, municipio_name)
    if not municipio_data or "veredas" not in municipio_data:
        logger.warning(f"❌ {municipio_name}: no tiene datos de veredas")
        return
    
    veredas_list = list(municipio_data["veredas"].keys())
    logger.info(f"🔍 {municipio_name}: {len(veredas_list)} veredas disponibles")
    
    # Buscar coincidencias parciales
    vereda_upper = str(vereda_name_shapefile).upper().strip()
    coincidencias = [v for v in veredas_list if vereda_upper in v.upper() or v.upper() in vereda_upper]
    
    if coincidencias:
        logger.info(f"🎯 Posibles: {coincidencias[:3]}")
    else:
        logger.warning(f"❌ Sin coincidencias para '{vereda_name_shapefile}'")

# ===== MANTENER FUNCIONES LEGACY PARA COMPATIBILIDAD =====

def calculate_urban_data_not_visualized(cobertura_data):
    """Calcula datos urbanos no visualizados."""
    total_urbano = {"poblacion": 0, "vacunados": 0, "municipios_con_urbano": 0, "cobertura": 0.0}
    
    if not cobertura_data or "municipios" not in cobertura_data:
        return total_urbano
    
    for municipio_data in cobertura_data["municipios"].values():
        urbano_data = municipio_data.get("urbano", {})
        if urbano_data.get("poblacion", 0) > 0:
            total_urbano["poblacion"] += urbano_data["poblacion"]
            total_urbano["vacunados"] += urbano_data["vacunados"]
            total_urbano["municipios_con_urbano"] += 1
    
    if total_urbano["poblacion"] > 0:
        total_urbano["cobertura"] = round((total_urbano["vacunados"] / total_urbano["poblacion"]) * 100, 1)
    
    return total_urbano